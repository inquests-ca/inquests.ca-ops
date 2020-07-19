import os
import sys
import argparse
import re
import datetime
from collections import defaultdict

import boto3
import botocore
import sqlalchemy
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook

from db.models import *


class Migrator:

    _AUTHORITY_TYPE_AUTHORITY = 'Authority'
    _AUTHORITY_TYPE_INQUEST = 'Inquest/Fatality Inquiry'

    def __init__(self, data_directory, document_files_directory, db_url, upload_documents):
        # TODO: run inquestsca.sql on startup.
        self._data_directory = data_directory
        self._document_files_directory = document_files_directory
        self._session_maker = self._init_session_maker(db_url)
        self._upload_documents = upload_documents
        self._s3_client, self._s3_client_url_generator = self._init_s3_clients()

        # Mappings from input data attributes to new IDs.
        self._authority_keyword_name_to_id = {}
        self._inquest_keyword_name_to_id = {}
        self._authority_serial_to_id = {}

        # Mappings from authority serial to various attributes.
        self._authority_serial_to_related = {}          # Tuple of related authorities and cited authorities.
        self._authority_serial_to_type = {}             # Type being either authority or inquest.
        self._authority_serial_to_name = {}
        self._authority_serial_to_keywords = {}
        self._authority_serial_to_primary_document = {}

        # This operation must be done first to satisfy FK constraints.
        self.populate_keywords()

        # This depends on the previous operation.
        self.populate_authorities_and_inquests()

        # These operations depend on all previous operations and are independent of each other.
        self.populate_authority_relationships()
        self.populate_authority_and_inquest_keywords()
        self.populate_documents()

        # Run checks to ensure data is valid.
        self.validate()

    def _init_session_maker(self, db_url):
        """Create session maker which will be used to initiate database connections."""
        engine = create_engine(db_url)
        Session = sessionmaker(bind=engine)
        return Session

    def _init_s3_clients(self):
        """Create S3 client used to interface the inquests-ca-resources S3 bucket."""
        session = boto3.Session(profile_name='migration')
        config = botocore.client.Config(signature_version=botocore.UNSIGNED)
        return session.client('s3'), session.client('s3', config=config)

    def _read_workbook(self, workbook):
        """Returns iterator for rows in given Excel file."""
        wb = load_workbook(os.path.join(self._data_directory, 'caspio_{}.xlsx'.format(workbook)))
        ws = wb.active

        # Start at 2nd row to ignore headers.
        return ws.iter_rows(min_row=2, values_only=True)

    def _is_valid_authority_type(self, authority_type):
        return authority_type in [self._AUTHORITY_TYPE_AUTHORITY, self._AUTHORITY_TYPE_INQUEST]

    def _format_as_id(self, name):
        """Formats string to an appropriate ID."""
        return (
            name
                .upper()
                .replace('-', '_')
                .replace(' ', '_')
                .replace('.', '_')
                .replace('/', '_')
        )

    def _nullable_to_string(self, string):
        """Done to satisfy NULL constraints."""
        if string is None:
            return ''
        return self._format_string(string)

    def _string_to_nullable(self, string):
        """If string is empty, return None"""
        if string is None or string.strip() == '':
            return None
        return self._format_string(string)

    def _format_string(self, string):
        """Format string before inserting into database. Currently only trims whitespace."""
        if string is None:
            return None
        return string.strip()

    def _format_date(self, date):
        """Format date string into SQL-compatible date."""
        if type(date) == datetime.datetime:
            return date
        elif date is None or date == '':
            return None
        elif re.match(r'\d{4}-\d{2}-\d{2}', date) is not None:
            # Date is already in valid format.
            return date
        else:
            match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', date)
            if match is not None and len(match.groups()) == 3:
                (month, day, year) = match.groups()
                return "{}-{}-{}".format(year, month, day)
            else:
                raise ValueError('Invalid date: {}'.format(date))

    def _get_year_from_date(self, date):
        """Get year from date; note that date may be one of many types or formats."""
        if type(date) == datetime.datetime:
            return str(date.year)
        elif date is None or date == '':
            return None
        elif re.match(r'\d{4}-\d{2}-\d{2}', date) is not None:
            return date[:4]
        elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', date) is not None:
            return date[-4:]
        else:
            raise ValueError('Invalid date: {}'.format(date))

    def _format_s3_key_segment(self, string):
        """Replaces certain characters in the given string to avoid the need for URL encoding."""
        if string is None:
            return 'MissingData'
        return re.sub(r'[^a-zA-Z0-9]+', '-', string).strip('-')

    def _jurisdiction_serial_to_id(self, serial):
        serial_to_id = {
            'CAN': 'CAN',
            'UK': 'UK',
            'US': 'US',
        }

        if serial in serial_to_id:
            return serial_to_id[serial]
        else:
            # In the default case, prepend CAN_ to serial to get ID.
            return 'CAN_{}'.format(serial)

    def _source_serial_to_id(self, serial):
        serial_to_id = {
            'CANLEG': 'CAN_LEG',
            'UKSenC': 'UK_SENC',
            'UKSC': 'UK_SC',
            'USSC': 'US_SC',
            'REF': 'REF',
        }

        if serial in serial_to_id:
            return serial_to_id[serial]
        else:
            # In the default case, prepend CAN_ to serial to get ID.
            return 'CAN_{}'.format(serial)

    def _upload_document_if_exists(self, name, date, source, serial, authority_serial):
        """Upload document file to S3 if one exists locally."""
        if serial is None:
            print(
                 '[WARNING] No serial for document: {}'
                .format(name)
            )
            return None

        # Get file path.
        directory = os.path.join(self._document_files_directory, serial.strip())
        documents = list(os.scandir(directory)) if os.path.isdir(directory) else []

        # Ensure there is exactly one file per document directory.
        if len(documents) != 1:
            print(
                 '[WARNING] Document {} has invalid number of files: {}.'
                .format(serial, len(documents))
            )
            return None

        file_path = documents[0].path

        year = self._get_year_from_date(date)
        source_id = self._source_serial_to_id(serial)

        # Generate S3 key for the given document with the form:
        # Documents/<source>/<year>/<authority name>/<document name>
        authority_name = self._authority_serial_to_name[authority_serial]
        key = '/'.join([
            'Documents',
            self._format_s3_key_segment(source_id),
            self._format_s3_key_segment(year),
            self._format_s3_key_segment(authority_name),
            self._format_s3_key_segment(name),
        ]) + '.pdf'

        bucket = 'inquests-ca-resources'

        # Currently no other way to get the object link with the Boto client.
        # See https://stackoverflow.com/a/48197877
        link = self._s3_client_url_generator.generate_presigned_url(
            'get_object',
            ExpiresIn=0,
            Params={
                'Bucket': bucket,
                'Key': key
            }
        )

        if not self._upload_documents:
            return link

        # Check if file exists to avoid unnecessary writes.
        try:
            obj = self._s3_client.get_object(
                Bucket=bucket,
                Key=key,
            )
            print(
                 '[DEBUG] Not uploading file since one already exists for document with ID: {}'
                .format(serial)
            )
        except self._s3_client.exceptions.NoSuchKey:
            self._s3_client.upload_file(
                file_path,
                bucket,
                key,
                ExtraArgs={
                    'ContentDisposition': 'inline',
                    'ContentType': 'application/pdf'
                },
            )
            print(
                 '[DEBUG] Successfully uploaded document with ID: {}, link: {}'
                .format(serial, link)
            )

        return link

    def populate_keywords(self):
        print('[INFO] Populating keywords.')

        session = self._session_maker()

        authority_categories = set()
        inquest_categories = set()

        for row in self._read_workbook('keywords'):
            rtype, rkeyword, rserial, rdescription = row

            if not self._is_valid_authority_type(rtype):
                print('[WARNING] Unknown authority type: {}'.format(rtype))
                continue

            # Most keywords are prefixed by categories (e.g., Cause-Fall from height -> Cause).
            # Otherwise, use 'General' as the default category.
            if '-' in rkeyword:
                category = (rkeyword.split('-', 1)[0])
            elif rkeyword == 'Evidence General':
                # Special case where - is not used.
                category = 'Evidence'
            else:
                category = 'General'

            # Create category if it does not exist.
            category_id = self._format_as_id(category)
            if rtype == self._AUTHORITY_TYPE_AUTHORITY and category_id not in authority_categories:
                session.add(AuthorityCategory(
                    authorityCategoryId=category_id,
                    name=self._format_string(category.title()),
                    description=rdescription
                ))
                authority_categories.add(category_id)
            elif rtype == self._AUTHORITY_TYPE_INQUEST and category_id not in inquest_categories:
                session.add(InquestCategory(
                    inquestCategoryId=category_id,
                    name=self._format_string(category.title()),
                    description=rdescription
                ))
                inquest_categories.add(category_id)
            session.flush()

            # Get keyword ID from keyword name (e.g., Cause-Fall from height -> CAUSE_FALL_FROM_HEIGHT).
            keyword_id = self._format_as_id(rkeyword)
            # Name keyword without category (e.g., Cause-Fall from height -> Fall from height)
            keyword_name = (rkeyword.split('-', 1)[1]) if '-' in rkeyword else rkeyword

            if rtype == self._AUTHORITY_TYPE_AUTHORITY:
                self._authority_keyword_name_to_id[rkeyword] = keyword_id
                model = AuthorityKeyword(
                    authorityKeywordId=keyword_id,
                    authorityCategoryId=category_id,
                    name=keyword_name,
                    description=None,
                )
            else:
                self._inquest_keyword_name_to_id[rkeyword] = keyword_id
                model = InquestKeyword(
                    inquestKeywordId=keyword_id,
                    inquestCategoryId=category_id,
                    name=keyword_name,
                    description=None,
                )
            session.add(model)

        session.commit()

    def populate_authorities_and_inquests(self):
        print('[INFO] Populating authorities and inquests.')

        session = self._session_maker()

        inquest_types = {
            'CONSTRUCTION',
            'CUSTODY_INMATE',
            'CUSTODY_POLICE',
            'DISCRETIONARY',
            'MINING',
            'PSYCHIATRIC_RESTRAINT'
        }
        death_manners = {
            'ACCIDENT',
            'HOMICIDE',
            'SUICIDE',
            'NATURAL',
            'UNDETERMINED',
        }

        for row in self._read_workbook('authorities'):
            (rserial, rname, _, rtype, rsynopsis, rkeywords, _, rquotes, rnotes, rprimary, _, _,
                roverview, _, _, rjurisdiction, _, _, rprimarydoc, _, rcited, rrelated, _, _, _,
                rlastname, rgivennames, rdeathdate, rcause, rinqtype, rpresidingofficer, rsex, rage,
                rstart, rend, _, _, _, _, _, rdeathmanner) = row

            if not self._is_valid_authority_type(rtype):
                print('[WARNING] Unknown authority type: {}'.format(rtype))
                continue

            if rtype == self._AUTHORITY_TYPE_AUTHORITY:
                authority = Authority(
                    isPrimary=rprimary,
                    name=self._format_string(rname),
                    overview=self._nullable_to_string(roverview),
                    synopsis=self._nullable_to_string(rsynopsis),
                    quotes=self._string_to_nullable(rquotes),
                    notes=self._string_to_nullable(rnotes)
                )
                session.add(authority)
                session.flush()
                authority_id = authority.authorityId
                self._authority_serial_to_related[authority_id] = (rcited, rrelated)
                self._authority_serial_to_primary_document[rserial] = rprimarydoc
            else:
                # Some inquests have their name prefixed with 'Inquest-'; this is redundant.
                if rname.startswith('Inquest-'):
                    rname = rname.replace('Inquest-', '', 1)

                # Parse description from inquest synopsis.
                match = re.search(r'Manner of Death: .*\s*((.|\n)*)', rsynopsis)
                if match is None:
                    # Assume that inquest synopsis only contains description of inquest.
                    synopsis = rsynopsis
                else:
                    synopsis = match.group(1)

                inquest = Inquest(
                    jurisdictionId=self._jurisdiction_serial_to_id(rjurisdiction),
                    isPrimary=rprimary,
                    name=self._format_string(rname),
                    overview=None,
                    synopsis=self._format_string(synopsis),
                    notes=self._string_to_nullable(rnotes),
                    presidingOfficer=self._nullable_to_string(rpresidingofficer),
                    start=self._format_date(rstart),
                    end=self._format_date(rend),
                    sittingDays=None,
                    exhibits=None,
                )
                session.add(inquest)
                session.flush()
                authority_id = inquest.inquestId

                # Validate inquest type.
                if rinqtype.startswith('Mandatory-'):
                    # An inquest is either 'Discretionary' or 'Mandatory-<reason>'; this makes 'Mandatory' redundant.
                    inquest_type = rinqtype.replace('Mandatory-', '')
                else:
                    inquest_type = rinqtype

                inquest_type_id = self._format_as_id(inquest_type)
                if inquest_type_id not in inquest_types:
                    print(
                        '[WARNING] Invalid inquest type: {} referenced by inquest with ID: {}. Defaulting to "OTHER".'
                        .format(inquest_type, rserial)
                    )
                    inquest_type_id = 'OTHER'

                # Validate manner of death.
                death_manner_id = self._format_as_id(rdeathmanner)
                if death_manner_id not in death_manners:
                    print(
                        '[WARNING] Invalid manner of death {} referenced by inquest with ID: {}. Defaulting to "OTHER".'
                        .format(rdeathmanner, rserial)
                    )
                    death_manner_id = 'OTHER'

                if rlastname == 'YOUTH' and self._string_to_nullable(rgivennames) is None:
                    # Names not available as by the Youth Criminal Justice Act.
                    last_name = None
                    given_names = None
                else:
                    last_name = self._format_string(rlastname.title())
                    given_names = self._format_string(rgivennames.title())

                deceased = Deceased(
                    inquestId=authority_id,
                    inquestTypeId=inquest_type_id,
                    deathMannerId=death_manner_id,
                    deathCause=self._format_string(rcause),
                    deathDate=self._format_date(rdeathdate),
                    lastName=last_name,
                    givenNames=given_names,
                    age=rage,
                    sex=(rsex if rsex != '?' else None)
                )
                session.add(deceased)

            self._authority_serial_to_type[rserial] = rtype
            self._authority_serial_to_name[rserial] = self._format_string(rname)
            self._authority_serial_to_keywords[rserial] = rkeywords
            self._authority_serial_to_id[rserial] = authority_id

        session.commit()

    def populate_authority_relationships(self):
        print('[INFO] Populating authority relationships.')

        session = self._session_maker()

        for (authority_id, (cited, related)) in self._authority_serial_to_related.items():
            # Map authority to its cited authorities and related authorities.
            if cited is not None:
                for authority_serial in cited.split('\n'):
                    if authority_serial == '':
                        continue

                    # Ignore references to authorities which do not exist.
                    if authority_serial not in self._authority_serial_to_id:
                        print(
                             '[WARNING] Invalid authority {} cited by authority with ID: {}'
                            .format(authority_serial, authority_id)
                        )
                        continue

                    # Ignore references to inquests.
                    if self._authority_serial_to_type[authority_serial] == self._AUTHORITY_TYPE_INQUEST:
                        print(
                             '[WARNING] Inquest {} cited by authority with ID: {}'
                            .format(authority_serial, authority_id)
                        )
                        continue

                    session.add(AuthorityCitations(
                        authorityId=authority_id,
                        citedAuthorityId=self._authority_serial_to_id[authority_serial],
                    ))

            if related is not None:
                for authority_serial in related.split('\n'):
                    if authority_serial == '':
                        continue

                    # Ignore references to authorities which do not exist.
                    if authority_serial not in self._authority_serial_to_id:
                        print(
                             '[WARNING] Invalid authority {} related to authority with ID: {}'
                            .format(authority_serial, authority_id)
                        )
                        continue

                    if self._authority_serial_to_type[authority_serial] == self._AUTHORITY_TYPE_INQUEST:
                        session.add(AuthorityInquests(
                            authorityId=authority_id,
                            inquestId=self._authority_serial_to_id[authority_serial],
                        ))
                    else:
                        session.add(AuthorityRelated(
                            authorityId=authority_id,
                            relatedAuthorityId=self._authority_serial_to_id[authority_serial],
                        ))

        session.commit()

    def populate_authority_and_inquest_keywords(self):
        print('[INFO] Populating authority and inquest keywords.')

        session = self._session_maker()

        for authority_serial, keywords in self._authority_serial_to_keywords.items():
            for keyword in keywords.split(','):
                if keyword == '' or keyword == 'zz_NotYetClassified':
                    continue

                if self._authority_serial_to_type[authority_serial] == self._AUTHORITY_TYPE_AUTHORITY:
                    if keyword not in self._authority_keyword_name_to_id:
                        print(
                             '[WARNING] Invalid keyword {} referenced by authority with ID: {}'
                            .format(keyword, authority_serial)
                        )
                        continue
                    session.add(AuthorityKeywords(
                        authorityId=self._authority_serial_to_id[authority_serial],
                        authorityKeywordId=self._authority_keyword_name_to_id[keyword],
                    ))
                else:
                    if keyword not in self._inquest_keyword_name_to_id:
                        print(
                             '[WARNING] Invalid keyword {} referenced by inquest with ID: {}'
                            .format(keyword, authority_serial)
                        )
                        continue
                    session.add(InquestKeywords(
                        inquestId=self._authority_serial_to_id[authority_serial],
                        inquestKeywordId=self._inquest_keyword_name_to_id[keyword],
                    ))

        session.commit()

    def populate_documents(self):
        print('[INFO] Populating authority and inquest documents.')

        session = self._session_maker()

        document_sources = set()

        for row in self._read_workbook('docs'):
            rauthorities, rserial, rshortname, rcitation, rdate, rlink, rlinktype, rsource = row

            if rlinktype != 'No Publish':
                # Create document source type (i.e., the location where the document is stored) if it does not exist.
                document_source_id = self._format_as_id(rlinktype)
                if document_source_id not in document_sources:
                    session.add(DocumentSource(
                        documentSourceId=document_source_id,
                        name=self._format_string(rlinktype),
                    ))
                    session.flush()
                    document_sources.add(document_source_id)

            # Ensure document references at least one authority.
            if not any(rauthorities.split('\n')):
                print(
                     '[WARNING] Document {} does not reference any authorities.'
                    .format(rserial)
                )

            # Map authority or inquest to its documents.
            for authority_serial in rauthorities.split('\n'):
                if authority_serial == '':
                    continue

                # Ignore references to authorities which do not exist.
                if authority_serial not in self._authority_serial_to_id:
                    print(
                         '[WARNING] Invalid authority {} referenced by document with ID: {}'
                        .format(authority_serial, rserial)
                    )
                    continue

                # Upload document to S3 if respective file exists locally.
                link = None
                if rlinktype == 'Inquests.ca':
                    if rlink is not None and len(rlink) != 0:
                        print(
                             '[WARNING] Document {} has source Inquests.ca and non-null link: {}'
                            .format(rserial, rlink)
                        )
                    s3_link = self._upload_document_if_exists(rshortname, rdate, rsource, rserial, authority_serial)
                    if s3_link is not None:
                        link = s3_link
                elif rlinktype == 'No Publish':
                    if rlink is not None and len(rlink) != 0:
                        print(
                             '[WARNING] Document {} has flag No Publish and non-null link: {}'
                            .format(rserial, rlink)
                        )
                else:
                    if rlink is not None and len(rlink) != 0:
                        link = rlink
                    else:
                        print(
                             '[WARNING] Document {} has null link.'
                            .format(rserial)
                        )

                if self._authority_serial_to_type[authority_serial] == self._AUTHORITY_TYPE_AUTHORITY:
                    authority_document = AuthorityDocument(
                        authorityId=self._authority_serial_to_id[authority_serial],
                        authorityDocumentTypeId=None,
                        sourceId=self._source_serial_to_id(rsource),
                        isPrimary=rcitation == self._authority_serial_to_primary_document[authority_serial],
                        name=self._format_string(rshortname),
                        citation=self._format_string(rcitation),
                        created=self._format_date(rdate),
                    )
                    session.add(authority_document)
                    session.flush()
                    if link is not None:
                        session.add(AuthorityDocumentLinks(
                            authorityDocumentId=authority_document.authorityDocumentId,
                            documentSourceId=document_source_id,
                            link=link,
                        ))
                else:
                    if rshortname.startswith('Inquest-'):
                        # Some inquest documents begin with 'Inquest-'; this is redundant.
                        document_name = rshortname.replace('Inquest-', '')
                    else:
                        document_name = rshortname

                    inquest_document = InquestDocument(
                        inquestId=self._authority_serial_to_id[authority_serial],
                        inquestDocumentTypeId=None,
                        name=self._format_string(document_name),
                        created=self._format_date(rdate),
                    )
                    session.add(inquest_document)
                    session.flush()
                    if link is not None:
                        session.add(InquestDocumentLinks(
                            inquestDocumentId=inquest_document.inquestDocumentId,
                            documentSourceId=document_source_id,
                            link=link,
                        ))

        session.commit()

    def validate(self):
        print('[INFO] Running SQL validation scripts.')

        session = self._session_maker()

        # Invert mapping of authority IDs to map new IDs to input IDs.
        authority_id_to_serial = {
            new_id: serial for serial, new_id in self._authority_serial_to_id.items()
            if self._authority_serial_to_type[serial] == self._AUTHORITY_TYPE_AUTHORITY
        }
        inquest_id_to_serial = {
            new_id: serial for serial, new_id in self._authority_serial_to_id.items()
            if self._authority_serial_to_type[serial] == self._AUTHORITY_TYPE_INQUEST
        }

        # Ensure each authority has exactly one primary document.
        query = sqlalchemy.text("""
            SELECT authority.authorityId, authorityDocument.isPrimary, COUNT(authority.authorityId) AS cnt
            FROM authority
            LEFT JOIN authorityDocument ON authority.authorityId = authorityDocument.authorityId AND authorityDocument.isPrimary = 1
            GROUP BY authority.authorityId, authorityDocument.isPrimary
            HAVING authorityDocument.isPrimary IS NULL OR cnt > 1;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            authority_id, has_primary, count = row
            if not has_primary:
                count = 0
            print(
                 '[WARNING] Authority {} has {} primary documents.'
                .format(authority_id_to_serial[authority_id], count)
            )

        # Ensure each inquest has at least one document.
        query = sqlalchemy.text("""
            SELECT inquest.inquestId
            FROM inquest
            LEFT JOIN inquestDocument ON inquest.inquestId = inquestDocument.inquestId
            WHERE inquestDocument.inquestId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            inquest_id = row[0]
            print(
                 '[WARNING] Inquest {} does not have any documents.'
                .format(inquest_id_to_serial[inquest_id], count)
            )

        # Ensure each authority has at least one keyword.
        query = sqlalchemy.text("""
            SELECT authority.authorityId
            FROM authority
            LEFT JOIN authorityKeywords ON authorityKeywords.authorityId = authority.authorityId
            WHERE authorityKeywords.authorityId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            authority_id = row[0]
            print(
                 '[WARNING] Authority {} does not have any keywords.'
                .format(authority_id_to_serial[authority_id], count)
            )

        # Ensure each inquest has at least one keyword.
        query = sqlalchemy.text("""
            SELECT inquest.inquestId
            FROM inquest
            LEFT JOIN inquestKeywords ON inquestKeywords.inquestId = inquest.inquestId
            WHERE inquestKeywords.inquestId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            inquest_id = row[0]
            print(
                 '[WARNING] Inquest {} does not have any keywords.'
                .format(inquest_id_to_serial[inquest_id], count)
            )

        # Ensure each inquest has at least one CAUSE keyword.
        query = sqlalchemy.text("""
            SELECT inquest.inquestId
            FROM inquestKeywords
            INNER JOIN inquestKeyword ON inquestKeyword.inquestKeywordId = inquestKeywords.inquestKeywordId AND inquestKeyword.inquestCategoryId = 'CAUSE'
            RIGHT JOIN inquest ON inquest.inquestId = inquestKeywords.inquestId
            WHERE inquestKeywords.inquestId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            inquest_id = row[0]
            print(
                 '[WARNING] Inquest {} does not have any CAUSE keywords.'
                .format(inquest_id_to_serial[inquest_id], count)
            )


if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('--data', help='Directory containing data to be processed')
    parser.add_argument('--documents', help='Directory containing documents')
    parser.add_argument('--db', help='Database URL')
    parser.add_argument('--upload', action='store_true', help='Whether to upload documents to AWS S3')

    args = parser.parse_args()

    Migrator(args.data, args.documents, args.db, args.upload)
