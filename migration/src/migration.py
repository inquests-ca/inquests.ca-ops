import os
import re

import sqlalchemy
from openpyxl import load_workbook

import models
import utils
from db import DatabaseClient
from logger import logger
from s3 import S3Client


class Migrator:

    _AUTHORITY_TYPE_AUTHORITY = 'Authority'
    _AUTHORITY_TYPE_INQUEST = 'Inquest/Fatality Inquiry'

    def __init__(self, data_directory, document_files_directory, db_url, upload_documents):
        self._data_directory = data_directory
        self._document_files_directory = document_files_directory
        self._upload_documents = upload_documents
        self._db_client = DatabaseClient(db_url)
        self._s3_client = S3Client(bucket='inquests-ca-resources')

        # Sets of authority and inquest keywords.
        self._authority_keyword_ids = set()
        self._inquest_keyword_ids = set()

        # Mappings from authority serial to various attributes.
        self._authority_serial_to_id = {}
        self._authority_serial_to_related = {}          # Tuple of related authorities and cited authorities.
        self._authority_serial_to_type = {}             # Type being either authority or inquest.
        self._authority_serial_to_name = {}
        self._authority_serial_to_primary_document = {}

    def run(self):
        # These operations must be done first to satisfy FK constraints.
        self.populate_sources()
        self.populate_keywords()

        # This depends on the previous operation.
        self.populate_authorities_and_inquests()

        # These operations depend on all previous operations and are independent of each other.
        self.populate_authority_relationships()
        self.populate_documents()

        # Run checks to ensure data is valid.
        self.validate()

    def _read_workbook(self, workbook):
        """Returns iterator for rows in given Excel file."""
        work_book = load_workbook(os.path.join(self._data_directory, 'caspio_{}.xlsx'.format(workbook)))
        work_sheet = work_book.active

        # Start at 2nd row to ignore headers.
        return work_sheet.iter_rows(min_row=2, values_only=True)

    def _is_valid_authority_type(self, authority_type):
        return authority_type in [self._AUTHORITY_TYPE_AUTHORITY, self._AUTHORITY_TYPE_INQUEST]

    def _is_federal_jurisdiction(self, jurisdiction_id):
        return jurisdiction_id in ['CAN', 'UK', 'US']

    def _jurisdiction_serial_to_id_and_category(self, serial):
        if serial in ['US', 'CAN']:
            return (serial, serial)
        elif serial == 'UKCOM':
            return ('UK', 'UK')
        elif serial == 'OTHER':
            return (None, None)
        else:
            # Otherwise jurisdiction is Canadian province or territory.
            return ('CAN_{}'.format(utils.format_as_id(serial)), 'CAN')

    def _source_serial_to_id(self, serial):
        serial_to_id = {
            'CANLEG': 'CAN_LEG',
            'CANPI': 'CAN_PI',
            'UKLEG': 'UK_LEG',
            'UKSenC': 'UK_SENC',
            'UKSC': 'UK_SC',
            'USOTH': 'US_OTHER',
            'USSC': 'US_SC',
            'REF': 'OTHER',
        }

        if serial in serial_to_id:
            return serial_to_id[serial]
        else:
            # In the default case, prepend CAN_ to serial to get ID.
            return 'CAN_{}'.format(serial)

    def _keyword_name_to_id(self, keyword):
        if utils.is_empty_string(keyword) or keyword == 'zz_NotYetClassified':
            return None

        return utils.format_as_id(keyword)

    def _upload_document_if_exists(self, name, date, source, serial, authority_serial):
        """Upload document file to S3 if one exists locally."""
        # Get file path.
        directory = os.path.join(self._document_files_directory, serial.strip())
        documents = list(os.scandir(directory)) if os.path.isdir(directory) else []

        # Ensure there is exactly one file per document directory.
        if len(documents) != 1:
            logger.warning('Document: %s has %d files.', serial, len(documents))
            return None

        file_path = documents[0].path

        year = utils.get_year_from_date(date)
        source_id = self._source_serial_to_id(source)

        # Generate S3 key for the given document with the form:
        # Documents/<source>/<year>/<authority name>/<document name>
        authority_name = self._authority_serial_to_name[authority_serial]
        key = self._s3_client.generate_s3_key(
            ['Documents', source_id, year, authority_name, name],
            'pdf'
        )

        link = self._s3_client.generate_object_url(key)

        if not self._upload_documents:
            return link

        # Check if file exists to avoid unnecessary writes.
        if self._s3_client.object_exists(key):
            logger.debug(
                'Document: %s will not be uploaded since one already exists.',
                serial
            )
        else:
            self._s3_client.upload_pdf(file_path, key)
            logger.debug('Document: %s successfully uploaded to: %s', serial, link)

        return link

    def populate_sources(self):
        logger.info('Populating sources.')

        session = self._db_client.get_session()

        for row in self._read_workbook('source'):
            rcode, rdescription, rjurisdiction, _, rrank = row

            if rcode.endswith('OTH'):
                code = 'OTHER'.join(rcode.rsplit('OTH', 1))
            elif rcode == 'REF':
                code = 'OTHER'
            else:
                code = rcode

            jurisdiction_serial = (rjurisdiction.split('-', 1)[0]).strip()
            jurisdiction_id, jurisdiction_category = self._jurisdiction_serial_to_id_and_category(jurisdiction_serial)
            if not jurisdiction_id:
                source_id = 'OTHER'
            elif self._is_federal_jurisdiction(jurisdiction_id) and code.startswith(jurisdiction_id):
                # Handles sources such as CANLEG and UKSenC.
                source_id = '{}_{}'.format(jurisdiction_category, code.replace(jurisdiction_id, '', 1))
            else:
                source_id = '{}_{}'.format(jurisdiction_category, code)

            rank = int((rrank.split('-', 1)[0]).strip())

            session.add(models.Source(
                sourceId=utils.format_as_id(source_id),
                jurisdictionId=jurisdiction_id,
                name=utils.format_string(rdescription),
                code=utils.format_as_id(code),
                rank=rank
            ))
            session.flush()

        session.commit()

    def populate_keywords(self):
        logger.info('Populating keywords.')

        session = self._db_client.get_session()

        authority_categories = {
            'EVIDENCE',
            'FACTOR',
            'INQUEST',
        }
        inquest_categories = {
            'CAUSE',
            'FACTOR',
            'INQUEST',
        }

        for row in self._read_workbook('keywords'):
            rtype, rkeyword, _, rdescription = row

            if not self._is_valid_authority_type(rtype):
                logger.warning('Keyword: "%s" has unknown authority type: "%s".', rkeyword, rtype)
                continue

            # Keywords are prefixed by category (e.g., Cause-Fall from height -> Cause).
            if '-' in rkeyword:
                category_id = utils.format_as_id(rkeyword.split('-', 1)[0])
            elif rkeyword == 'Evidence General':
                # Special case where - is not used.
                category_id = 'EVIDENCE'

            keyword_id = self._keyword_name_to_id(rkeyword)
            if keyword_id is None:
                logger.warning('Keyword: "%s" is invalid.', rkeyword)

            # Name keyword without category (e.g., Cause-Fall from height -> Fall from height)
            keyword_name = (rkeyword.split('-', 1)[1]) if '-' in rkeyword else rkeyword

            if rtype == self._AUTHORITY_TYPE_AUTHORITY:
                if category_id not in authority_categories:
                    logger.warning(
                        'Keyword: "%s" has invalid authority category: "%s".',
                        rkeyword, category_id
                    )
                    continue
                self._authority_keyword_ids.add(keyword_id)
                model = models.AuthorityKeyword(
                    authorityKeywordId=keyword_id,
                    authorityCategoryId=category_id,
                    name=keyword_name,
                    description=rdescription,
                )
            else:
                if category_id not in inquest_categories:
                    logger.warning(
                        'Keyword: "%s" has invalid inquest category: "%s".',
                        rkeyword, category_id
                    )
                    continue
                self._inquest_keyword_ids.add(keyword_id)
                model = models.InquestKeyword(
                    inquestKeywordId=keyword_id,
                    inquestCategoryId=category_id,
                    name=keyword_name,
                    description=rdescription,
                )
            session.add(model)

        session.commit()

    def populate_authorities_and_inquests(self):
        logger.info('Populating authorities and inquests.')

        session = self._db_client.get_session()

        # Separate authorities by type and sort by export ID
        rauthorities = []
        rinquests = []
        for row in sorted(self._read_workbook('authorities'), key=lambda row: row[-1]):
            rserial = row[0]
            rtype = row[3]

            if rtype == self._AUTHORITY_TYPE_AUTHORITY:
                rauthorities.append(row)
            elif rtype == self._AUTHORITY_TYPE_INQUEST:
                rinquests.append(row)
            else:
                logger.warning(
                    'Authority: %s has unknown authority type: "%s".',
                    rserial, rtype
                )
                continue

        for rauthority in rauthorities:
            (rserial, rname, _, _, rsynopsis, rkeywords, rtags, rquotes, rnotes, rprimary, _, _,
                roverview, _, _, _, _, _, rprimarydoc, _, rcited, rrelated, _, _, _, _, _, _, _, _,
                _, _, _, _, _, _, _, _, _, _, _, rexport) = rauthority

            authority_id = self._create_authority(
                session, rserial, rname, rsynopsis, rquotes, rnotes, rprimary, roverview, rexport
            )

            self._authority_serial_to_type[rserial] = self._AUTHORITY_TYPE_AUTHORITY
            self._authority_serial_to_name[rserial] = utils.format_string(rname)
            self._authority_serial_to_id[rserial] = authority_id
            self._authority_serial_to_primary_document[rserial] = rprimarydoc
            self._authority_serial_to_related[rserial] = (rcited, rrelated)

            self._create_authority_keywords(session, authority_id, rserial, rkeywords)
            self._create_authority_tags(session, authority_id, rserial, rtags)

        session.commit()

        for rinquest in rinquests:
            (rserial, rname, _, _, rsynopsis, rkeywords, rtags, _, rnotes, rprimary, _, _, _, _, _,
                rjurisdiction, _, _, _, _, rcited, rrelated, _, _, _, rlastname, rgivennames,
                rdeathdate, rcause, rinqtype, rpresidingofficer, rsex, rage, rstart, rend, _, _, _,
                _, _, rdeathmanner, rexport) = rinquest

            if not utils.is_empty_string(rcited):
                logger.warning('Inquest: %s has citations, ignoring.', rserial)
            if not utils.is_empty_string(rrelated):
                logger.warning('Inquest: %s has related authorities, ignoring.', rserial)

            inquest_id = self._create_inquest(
                session, rserial, rname, rsynopsis, rnotes, rprimary, rjurisdiction,
                rpresidingofficer, rstart, rend, rexport
            )

            self._authority_serial_to_type[rserial] = self._AUTHORITY_TYPE_INQUEST
            self._authority_serial_to_name[rserial] = utils.format_string(rname)
            self._authority_serial_to_id[rserial] = inquest_id

            self._create_inquest_deceased(
                session, inquest_id, rserial, rlastname, rgivennames, rdeathdate, rcause, rinqtype,
                rsex, rage, rdeathmanner
            )
            self._create_inquest_keywords(session, inquest_id, rserial, rkeywords)
            self._create_inquest_tags(session, inquest_id, rserial, rtags)

        session.commit()

    def _create_authority(
            self, session, rserial, rname, rsynopsis, rquotes, rnotes, rprimary, roverview, rexport
        ):
        authority = models.Authority(
            isPrimary=rprimary,
            name=utils.format_string(rname),
            overview=utils.nullable_to_string(roverview),
            synopsis=utils.nullable_to_string(rsynopsis),
            quotes=utils.string_to_nullable(rquotes),
            notes=utils.string_to_nullable(rnotes)
        )
        session.add(authority)
        session.flush()
        authority_id = authority.authorityId
        assert authority_id == rexport,\
            "Autogenerated authority ID should match export ID for authority with ID: {}".format(rserial)

        return authority_id

    def _create_authority_keywords(self, session, authority_id, rserial, rkeywords):
        if utils.is_empty_string(rkeywords):
            return

        for keyword in rkeywords.split(','):
            keyword_id = self._keyword_name_to_id(keyword)
            if keyword_id is None:
                continue

            if keyword_id not in self._authority_keyword_ids:
                logger.warning(
                    'Authority: %s references invalid keyword: "%s".',
                    rserial, keyword
                )
                continue

            session.add(models.AuthorityKeywords(
                authorityId=authority_id,
                authorityKeywordId=keyword_id,
            ))

    def _create_authority_tags(self, session, authority_id, rserial, rtags):
        if utils.is_empty_string(rtags):
            return

        tags = set()

        for tag in re.split(r'[,\n]', rtags):
            if utils.is_empty_string(tag):
                continue

            tag = utils.format_as_keyword(tag)

            # Note that MySQL is case-insensitive for the UNIQUE constraint.
            if tag.lower() in tags:
                continue
            tags.add(tag.lower())

            session.add(models.AuthorityTags(
                authorityId=authority_id,
                tag=tag,
            ))

    def _create_inquest(
            self, session, rserial, rname, rsynopsis, rnotes, rprimary, rjurisdiction,
            rpresidingofficer, rstart, rend, rexport
        ):
        # Some inquests have their name prefixed with 'Inquest-'; this is redundant.
        if rname.startswith('Inquest-'):
            rname = rname.replace('Inquest-', '', 1)

        # Parse description from inquest synopsis.
        match = re.search(r'Manner of Death: .*\s*((.|\n)*)', rsynopsis)
        if match is None:
            # Assume that inquest synopsis only contains description of inquest.
            synopsis = rsynopsis
        else:
            synopsis = match.group(1)

        inquest = models.Inquest(
            jurisdictionId=self._jurisdiction_serial_to_id_and_category(rjurisdiction)[0],
            isPrimary=rprimary,
            name=utils.format_string(rname),
            overview=None,
            synopsis=utils.format_string(synopsis),
            notes=utils.string_to_nullable(rnotes),
            presidingOfficer=utils.nullable_to_string(rpresidingofficer),
            start=utils.format_date(rstart),
            end=utils.format_date(rend),
            sittingDays=None,
            exhibits=None,
        )
        session.add(inquest)
        session.flush()
        inquest_id = inquest.inquestId
        assert inquest_id == rexport,\
            "Autogenerated inquest ID should match export ID for inquest with ID: {}".format(rserial)

        return inquest_id

    def _create_inquest_deceased(
            self, session, inquest_id, rserial, rlastname, rgivennames, rdeathdate, rcause,
            rinqtype, rsex, rage, rdeathmanner
        ):
        inquest_types = {
            'CONSTRUCTION',
            'CUSTODY_INMATE',
            'CUSTODY_POLICE',
            'DISCRETIONARY',
            'MINING',
            'PSYCHIATRIC_RESTRAINT'
        }
        death_manners = {
            'ACCIDENT',
            'HOMICIDE',
            'SUICIDE',
            'NATURAL',
            'UNDETERMINED',
        }

        # Validate inquest type.
        if rinqtype.startswith('Mandatory-'):
            # An inquest is either 'Discretionary' or 'Mandatory-<reason>'; this makes 'Mandatory' redundant.
            inquest_type = rinqtype.replace('Mandatory-', '')
        else:
            inquest_type = rinqtype

        inquest_type_id = utils.format_as_id(inquest_type)
        if inquest_type_id not in inquest_types:
            logger.warning(
                'Inquest: %s has invalid inquest type: "%s". Defaulting to "OTHER".',
                rserial, inquest_type
            )
            inquest_type_id = 'OTHER'

        # Validate manner of death.
        death_manner_id = utils.format_as_id(rdeathmanner)
        if death_manner_id not in death_manners:
            logger.warning(
                'Inquest: %s has invalid manner of death: "%s". Defaulting to "OTHER".',
                rserial, rdeathmanner
            )
            death_manner_id = 'OTHER'

        if rlastname == 'YOUTH' and utils.string_to_nullable(rgivennames) is None:
            # Names not available as by the Youth Criminal Justice Act.
            last_name = None
            given_names = None
        else:
            last_name = utils.format_string(rlastname.title())
            given_names = utils.format_string(rgivennames.title())

        deceased = models.Deceased(
            inquestId=inquest_id,
            inquestTypeId=inquest_type_id,
            deathMannerId=death_manner_id,
            deathCause=utils.format_string(rcause),
            deathDate=utils.format_date(rdeathdate),
            lastName=last_name,
            givenNames=given_names,
            age=rage,
            sex=(rsex if rsex != '?' else None)
        )
        session.add(deceased)

    def _create_inquest_keywords(self, session, inquest_id, rserial, rkeywords):
        if utils.is_empty_string(rkeywords):
            return

        for keyword in rkeywords.split(','):
            keyword_id = self._keyword_name_to_id(keyword)
            if keyword_id is None:
                continue

            if keyword_id not in self._inquest_keyword_ids:
                logger.warning(
                    'Inquest: %s references invalid keyword "%s".',
                    rserial, keyword
                )
                continue

            session.add(models.InquestKeywords(
                inquestId=inquest_id,
                inquestKeywordId=keyword_id,
            ))

    def _create_inquest_tags(self, session, inquest_id, rserial, rtags):
        if utils.is_empty_string(rtags):
            return

        tags = set()

        for tag in re.split(r'[,\n]', rtags):
            if utils.is_empty_string(tag):
                continue

            tag = utils.format_as_keyword(tag)

            # Note that MySQL is case-insensitive for the UNIQUE constraint.
            if tag.lower() in tags:
                continue
            tags.add(tag.lower())

            session.add(models.InquestTags(
                inquestId=inquest_id,
                tag=tag,
            ))

    def populate_authority_relationships(self):
        logger.info('Populating authority relationships.')

        session = self._db_client.get_session()

        for (serial, (cited, related)) in self._authority_serial_to_related.items():
            # Map authority to its cited authorities and related authorities.
            if cited is not None:
                for cited_serial in cited.split('\n'):
                    if utils.is_empty_string(cited_serial):
                        continue

                    # Ignore references to authorities which do not exist.
                    if cited_serial not in self._authority_serial_to_id:
                        logger.warning(
                            'Authority: %s cites invalid authority: %s',
                            serial, cited_serial
                        )
                        continue

                    # Ignore references to inquests.
                    if self._authority_serial_to_type[cited_serial] == self._AUTHORITY_TYPE_INQUEST:
                        logger.warning(
                            'Authority: %s cites inquest: %s',
                            serial, cited_serial
                        )
                        continue

                    session.add(models.AuthorityCitations(
                        authorityId=self._authority_serial_to_id[serial],
                        citedAuthorityId=self._authority_serial_to_id[cited_serial],
                    ))

            if related is not None:
                for related_serial in related.split('\n'):
                    if utils.is_empty_string(related_serial):
                        continue

                    # Ignore references to authorities which do not exist.
                    if related_serial not in self._authority_serial_to_id:
                        logger.warning(
                            'Authority: %s is related to invalid authority: %s',
                            serial, related_serial
                        )
                        continue

                    if self._authority_serial_to_type[related_serial] == self._AUTHORITY_TYPE_INQUEST:
                        session.add(models.AuthorityInquests(
                            authorityId=self._authority_serial_to_id[serial],
                            inquestId=self._authority_serial_to_id[related_serial],
                        ))
                    else:
                        session.add(models.AuthorityRelated(
                            authorityId=self._authority_serial_to_id[serial],
                            relatedAuthorityId=self._authority_serial_to_id[related_serial],
                        ))

        session.commit()

    def populate_documents(self):
        logger.info('Populating authority and inquest documents.')

        session = self._db_client.get_session()

        document_sources = set()

        for row in self._read_workbook('docs'):
            rauthorities, rserial, rshortname, rcitation, rdate, rlink, rlinktype, rsource = row

            if rlinktype.lower() != 'no publish':
                # Create document source type (i.e., the location where the document is stored) if it does not exist.
                document_source_id = utils.format_as_id(rlinktype)
                if document_source_id not in document_sources:
                    session.add(models.DocumentSource(
                        documentSourceId=document_source_id,
                        name=utils.format_string(rlinktype),
                    ))
                    session.flush()
                    document_sources.add(document_source_id)

            # Ensure document references at least one authority.
            if not any(rauthorities.split('\n')):
                logger.warning('Document: %s does not reference any authorities.', rserial)

            # Map authority or inquest to its documents.
            for authority_serial in rauthorities.split('\n'):
                if utils.is_empty_string(authority_serial):
                    continue

                # Ignore references to authorities which do not exist.
                if authority_serial not in self._authority_serial_to_id:
                    logger.warning(
                        'Document: %s references invalid authority: %s',
                        rserial, authority_serial
                    )
                    continue

                # Upload document to S3 if respective file exists locally.
                link = None
                if rlinktype.lower() == 'inquests.ca':
                    if not utils.is_empty_string(rlink):
                        logger.warning(
                            'Document: %s has source Inquests.ca and non-null link: %s',
                            rserial, rlink
                        )
                    s3_link = self._upload_document_if_exists(rshortname, rdate, rsource, rserial, authority_serial)
                    if s3_link is not None:
                        link = s3_link
                elif rlinktype.lower() == 'no publish':
                    if not utils.is_empty_string(rlink):
                        logger.warning(
                            'Document: %s has flag No Publish and non-null link: %s',
                            rserial, rlink
                        )
                    else:
                        logger.debug('Document: %s has "No Publish" and will not be uploaded.', rserial)
                else:
                    if not utils.is_empty_string(rlink):
                        link = rlink
                    else:
                        logger.warning('Document: %s has null link.', rserial)

                if self._authority_serial_to_type[authority_serial] == self._AUTHORITY_TYPE_AUTHORITY:
                    authority_document = models.AuthorityDocument(
                        authorityId=self._authority_serial_to_id[authority_serial],
                        authorityDocumentTypeId=None,
                        sourceId=self._source_serial_to_id(rsource),
                        isPrimary=rcitation == self._authority_serial_to_primary_document[authority_serial],
                        name=utils.format_string(rshortname),
                        citation=utils.format_string(rcitation),
                        created=utils.format_date(rdate),
                    )
                    session.add(authority_document)
                    session.flush()
                    if link is not None:
                        session.add(models.AuthorityDocumentLinks(
                            authorityDocumentId=authority_document.authorityDocumentId,
                            documentSourceId=document_source_id,
                            link=link,
                        ))
                else:
                    if rshortname.startswith('Inquest-'):
                        # Some inquest documents begin with 'Inquest-'; this is redundant.
                        document_name = rshortname.replace('Inquest-', '')
                    else:
                        document_name = rshortname

                    inquest_document = models.InquestDocument(
                        inquestId=self._authority_serial_to_id[authority_serial],
                        inquestDocumentTypeId=None,
                        name=utils.format_string(document_name),
                        created=utils.format_date(rdate),
                    )
                    session.add(inquest_document)
                    session.flush()
                    if link is not None:
                        session.add(models.InquestDocumentLinks(
                            inquestDocumentId=inquest_document.inquestDocumentId,
                            documentSourceId=document_source_id,
                            link=link,
                        ))

        session.commit()

    def validate(self):
        logger.info('Running SQL validation scripts.')

        session = self._db_client.get_session()

        # Invert mapping of authority IDs to map new IDs to input IDs.
        authority_id_to_serial = {
            new_id: serial for serial, new_id in self._authority_serial_to_id.items()
            if self._authority_serial_to_type[serial] == self._AUTHORITY_TYPE_AUTHORITY
        }
        inquest_id_to_serial = {
            new_id: serial for serial, new_id in self._authority_serial_to_id.items()
            if self._authority_serial_to_type[serial] == self._AUTHORITY_TYPE_INQUEST
        }

        # Ensure each authority has exactly one primary document.
        query = sqlalchemy.text("""
            SELECT authority.authorityId, authorityDocument.isPrimary, COUNT(authority.authorityId) AS cnt
            FROM authority
            LEFT JOIN authorityDocument ON authority.authorityId = authorityDocument.authorityId AND authorityDocument.isPrimary = 1
            GROUP BY authority.authorityId, authorityDocument.isPrimary
            HAVING authorityDocument.isPrimary IS NULL OR cnt > 1;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            authority_id, has_primary, count = row
            if not has_primary:
                count = 0
            logger.warning(
                'Authority: %s has %d primary documents.',
                authority_id_to_serial[authority_id], count
            )

        # Ensure each inquest has at least one document.
        query = sqlalchemy.text("""
            SELECT inquest.inquestId
            FROM inquest
            LEFT JOIN inquestDocument ON inquest.inquestId = inquestDocument.inquestId
            WHERE inquestDocument.inquestId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            inquest_id = row[0]
            logger.warning(
                'Inquest: %s does not have any documents.',
                inquest_id_to_serial[inquest_id]
            )

        # Ensure each authority has at least one keyword.
        query = sqlalchemy.text("""
            SELECT authority.authorityId
            FROM authority
            LEFT JOIN authorityKeywords ON authorityKeywords.authorityId = authority.authorityId
            WHERE authorityKeywords.authorityId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            authority_id = row[0]
            logger.warning(
                'Authority: %s does not have any keywords.',
                authority_id_to_serial[authority_id]
            )

        # Ensure each inquest has at least one keyword.
        query = sqlalchemy.text("""
            SELECT inquest.inquestId
            FROM inquest
            LEFT JOIN inquestKeywords ON inquestKeywords.inquestId = inquest.inquestId
            WHERE inquestKeywords.inquestId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            inquest_id = row[0]
            logger.warning(
                'Inquest: %s does not have any keywords.',
                inquest_id_to_serial[inquest_id]
            )

        # Ensure each inquest has at least one CAUSE keyword.
        query = sqlalchemy.text("""
            SELECT inquest.inquestId
            FROM inquestKeywords
            INNER JOIN inquestKeyword ON inquestKeyword.inquestKeywordId = inquestKeywords.inquestKeywordId AND inquestKeyword.inquestCategoryId = 'CAUSE'
            RIGHT JOIN inquest ON inquest.inquestId = inquestKeywords.inquestId
            WHERE inquestKeywords.inquestId IS NULL;
        """)
        rows = session.execute(query).fetchall()

        for row in rows:
            inquest_id = row[0]
            logger.warning(
                'Inquest: %s does not have any CAUSE keywords.',
                inquest_id_to_serial[inquest_id]
            )
